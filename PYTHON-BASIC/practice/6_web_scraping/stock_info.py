"""
There is a list of most active Stocks on Yahoo Finance https://finance.yahoo.com/most-active.
You need to compose several sheets based on data about companies from this list.
To fetch data from webpage you can use requests lib. To parse html you can use beautiful soup lib or lxml.
Sheets which are needed:
1. 5 stocks with most youngest CEOs and print sheet to output. You can find CEO info in Profile tab of concrete stock.
    Sheet's fields: Name, Code, Country, Employees, CEO Name, CEO Year Born.
2. 10 stocks with best 52-Week Change. 52-Week Change placed on Statistics tab.
    Sheet's fields: Name, Code, 52-Week Change, Total Cash
3. 10 largest holds of Blackrock Inc. You can find related info on the Holders tab.
    Blackrock Inc is an investment management corporation.
    Sheet's fields: Name, Code, Shares, Date Reported, % Out, Value.
    All fields except first two should be taken from Holders tab.


Example for the first sheet (you need to use same sheet format):
==================================== 5 stocks with most youngest CEOs ===================================
| Name        | Code | Country       | Employees | CEO Name                             | CEO Year Born |
---------------------------------------------------------------------------------------------------------
| Pfizer Inc. | PFE  | United States | 78500     | Dr. Albert Bourla D.V.M., DVM, Ph.D. | 1962          |
...

About sheet format:
- sheet title should be aligned to center
- all columns should be aligned to the left
- empty line after sheet

Write at least 2 tests on your choose.
Links:
    - requests docs: https://docs.python-requests.org/en/latest/
    - beautiful soup docs: https://www.crummy.com/software/BeautifulSoup/bs4/doc/
    - lxml docs: https://lxml.de/
"""
# -------------------------------------code for ceo profile-----
import time
import requests
from bs4 import BeautifulSoup
import openpyxl
excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = "CEO"
print(excel.sheetnames)
sheet.append(['Name', 'Title', 'Pay', 'Exercised', 'Year_Born'])
headers = {
    'Accept': "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    'User-Agent': "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:133.0) Gecko/20100101 Firefox/133.0"
}
url = "https://finance.yahoo.com/quote/BBCP/profile/"
response = requests.get(url, headers=headers)
time.sleep(2)
soup = BeautifulSoup(response.content, 'html.parser')
table = soup.find("table", class_="yf-mj92za")
rows = table.find_all("tr")
for row in rows:
    cols = row.find_all("td")
    if cols:
        cols_text = [col.get_text().strip() for col in cols]
        print(" ".join(cols_text))
        sheet.append(cols_text)
excel.save("CEO_Profile.xlsx")
print("Data saved to CEO_Profile.xlsx")





# ----------------------------# code for stocks with best52-week change
import requests
from bs4 import BeautifulSoup
import openpyxl
import time

# Setup Excel
excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = "52-Week Change"
print(excel.sheetnames)
sheet.append(['Name', 'Code', '52-Week Change', 'Total Cash'])

# Headers
headers = {
    'User-Agent': "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:133.0) Gecko/20100101 Firefox/133.0"
}

# Step 1: Get most active stocks
url = "https://finance.yahoo.com/most-active/"
response = requests.get(url, headers=headers)
soup = BeautifulSoup(response.content, 'html.parser')

symbols = []
for a in soup.find_all('a', {'data-test': 'quoteLink'}):
    text = a.text.strip()
    if text not in symbols:
        symbols.append(text)
    if len(symbols) >= 20:
        break

# Step 2: For each symbol, go to stats page
data_rows = []
for symbol in symbols:
    stats_url = f"https://finance.yahoo.com/quote/{symbol}/key-statistics"
    stats_page = requests.get(stats_url, headers=headers)
    stats_soup = BeautifulSoup(stats_page.content, 'html.parser')

    name_tag = stats_soup.find("h1")
    name = name_tag.text.strip() if name_tag else "N/A"

    change = "N/A"
    cash = "N/A"

    try:
        tables = stats_soup.find_all("table")
        for table in tables:
            for row in table.find_all("tr"):
                cells = row.find_all("td")
                if len(cells) == 2:
                    label = cells[0].text.strip()
                    value = cells[1].text.strip()

                    if "52-Week Change" in label:
                        change = value
                    elif "Total Cash" in label:
                        cash = value
    except:
        pass

    data_rows.append((name, symbol, change, cash))
    time.sleep(1)

# Sort by 52-week change value (remove % and convert to float)
def clean_percent(val):
    try:
        return float(val.strip('%'))
    except:
        return float('-inf')

sorted_data = sorted(data_rows, key=lambda x: clean_percent(x[2]), reverse=True)[:10]

# Write to Excel
for row in sorted_data:
    sheet.append(row)

excel.save("Stocks_52WeekChange.xlsx")
print("Data saved to Stocks_52WeekChange.xlsx")




# ----------------------------# code for largest holdings of blackrock Inc.
import requests
from bs4 import BeautifulSoup
import openpyxl
import time

# Setup Excel
excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = "BlackRock Holdings"
print(excel.sheetnames)
sheet.append(['Name', 'Code', 'Shares', 'Date Reported', '% Out', 'Value'])

# Target: BlackRock Inc.
symbol = "BLK"
headers = {
    'User-Agent': "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:133.0) Gecko/20100101 Firefox/133.0"
}
holders_url = f"https://finance.yahoo.com/quote/{symbol}/holders"
page = requests.get(holders_url, headers=headers)
soup = BeautifulSoup(page.content, 'html.parser')

# Step 1: Locate "Top Institutional Holders" table
tables = soup.find_all("table")
target_table = None
for table in tables:
    if "Shares" in table.text and "% Out" in table.text:
        target_table = table
        break

if not target_table:
    print(" Could not find the holdings table.")
    exit()

# Step 2: Parse rows
rows = target_table.find_all("tr")[1:11]  # Skip header, get top 10
for row in rows:
    cols = row.find_all("td")
    if len(cols) >= 5:
        name = cols[0].text.strip()
        shares = cols[1].text.strip()
        date_reported = cols[2].text.strip()
        percent_out = cols[3].text.strip()
        value = cols[4].text.strip()
        sheet.append([name, symbol, shares, date_reported, percent_out, value])

excel.save("BlackRock_Holdings.xlsx")
print("Data saved to BlackRock_Holdings.xlsx")



# -----------------------test_task.py

import unittest
import openpyxl
import os

class TestExcelOutputs(unittest.TestCase):

    def test_ceo_file_created_and_has_data(self):
        file = "CEO_Profile.xlsx"
        self.assertTrue(os.path.exists(file), "CEO Excel file was not created.")
        
        wb = openpyxl.load_workbook(file)
        sheet = wb["CEO"]
        self.assertGreater(sheet.max_row, 1, "CEO sheet is empty (no data rows).")
        self.assertEqual(sheet.cell(row=1, column=1).value, "Name", "Header mismatch in CEO sheet.")

    # def test_52week_change_data(self):
    #     file = "Stocks_52WeekChange.xlsx"
    #     self.assertTrue(os.path.exists(file), "52-week Excel file was not created.")

    #     wb = openpyxl.load_workbook(file)
    #     sheet = wb.active
    #     self.assertEqual(sheet.title, "52-Week Change", "Sheet name is incorrect.")
    #     self.assertGreater(sheet.max_row, 1, "52-week sheet has no data.")

    def test_blackrock_holdings_sheet(self):
        file = "BlackRock_Holdings.xlsx"  # or "BlackRock_Holdings_Selenium.xlsx" if using Selenium
        self.assertTrue(os.path.exists(file), "BlackRock Excel file was not created.")

        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        self.assertIn(sheet.title, ["BlackRock Holdings"], "Incorrect or missing sheet.")
        self.assertGreaterEqual(sheet.max_row, 2, "Holdings sheet should have at least one data row.")

if __name__ == "__main__":
    unittest.main()



